'''
Excel工具类
传入Excel目录和sheet名
返回数据list
'''
from openpyxl import load_workbook
from FmApp.Utils.ConfigUtil import ConfigUtil

class ExcelUtil:
    def __init__(self, filePath,sheetName):
        self.filePath = filePath
        self.sheetName = sheetName

    def getExcelData(self):
        mode = ConfigUtil().doConfig("D:\Python\Study\FmApp\Configs\case.config","MODE","excute_case")
        wb = load_workbook(self.filePath)
        sheet = wb[self.sheetName]
        caseList = []
        for i in range(2, sheet.max_row + 1):
            dict = {}
            for j in range(1, sheet.max_column + 1):
                key = sheet.cell(1, j).value
                values = sheet.cell(i, j).value
                dict[key] = values
            caseList.append(dict)

        # 将要执行的用例返回
        if mode.lower() == "all":
            final_caseList = caseList
        else:
            final_caseList = []
            for i in caseList:
                if i["id"] in eval(mode):
                    final_caseList.append(i)
        return final_caseList

if __name__ == '__main__':
    list = ExcelUtil("D:\Python\Study\FmApp\TestData\\test_data.xlsx","login").getExcelData()
    print(list)

