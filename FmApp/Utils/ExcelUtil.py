"""
Excel工具类
传入Excel目录和sheet名
返回数据list
"""
from openpyxl import load_workbook
from FmApp.Utils.ConfigUtil import ConfigUtil

class ExcelUtil:
    def __init__(self, filePath,sheetName,resultList = None):
        self.filePath = filePath
        self.sheetName = sheetName
        self.resultList = resultList

    # 读取Excel测试数据
    def getExcelData(self):
        mode = ConfigUtil().doConfig(r"..\Configs\case.config","MODE","excute_case")
        wb = load_workbook(self.filePath)
        sheet = wb[self.sheetName]
        caseList = []
        for i in range(2, sheet.max_row + 1):
            dict = {}
            for j in range(1, sheet.max_column + 1):
                key = sheet.cell(1, j).value
                values = sheet.cell(i, j).value
                dict[key] = values
            caseList.append(dict)

        # 根据配置文件，将要执行的用例返回
        if mode.lower() == "all":
            final_caseList = caseList
        else:
            final_caseList = []
            for i in caseList:
                if i["id"] in eval(mode):
                    final_caseList.append(i)
        return final_caseList

    # 将测试结果写回Excel中
    def setExcelData(self):
        wb = load_workbook(self.filePath)
        sheet = wb[self.sheetName]
        for item in self.resultList:
            id = item["id"]
            sheet.cell(id+1,6).value = ""
            sheet.cell(id+1,6).value = str(item["actual_result"]) # 注意这里要转成str类型
            sheet.cell(id+1,7).value = ""
            sheet.cell(id+1,7).value = item["is_success"]
        wb.save(self.filePath)



if __name__ == '__main__':
    list = ExcelUtil(r"..\TestData\test_data.xlsx","login").getExcelData()
    print(list)
    # list = [{"id":"1","response_json":"{\"code\":\"200\"}","is_success":"True"},{"id":"2","response_json":"{\"code\":\"200300\"}","is_success":"False"}]
    # ExcelUtil("..\TestData\\test_data.xlsx","login",list).setExcelData()



