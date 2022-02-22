from openpyxl import load_workbook

# 打开Excel
wb = load_workbook("test_excel.xlsx")

# 定位表单
sheet = wb["Sheet1"]

# 定位单元格,获取值
res = sheet.cell(1,1).value

print(res)

print(sheet.max_row)
print(sheet.max_column)